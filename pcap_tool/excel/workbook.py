"""Excel (.xlsx) workbook generation for all extracted findings/intel."""

import re as _re
import ipaddress
from collections import defaultdict

from ..constants import CLEARTEXT_PROTOS, LATERAL_PROTOS, SUSPICIOUS_PORTS, WELL_KNOWN

def generate_xlsx(rows, nodes, edges, findings, cleartext_hits, banner_hits, tls_sessions, dns_events, wifi_data, output_path, certificates=None):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[!] openpyxl not installed — skipping XLSX.  pip install openpyxl")
        return False

    wb = Workbook()

    # ── Common styles ──────────────────────────────────────────────────────────
    def hdr_style(fill_hex="1F4E79"):
        return dict(
            font  =Font(name="Arial", bold=True, color="FFFFFF", size=10),
            fill  =PatternFill("solid", fgColor=fill_hex),
            align =Alignment(horizontal="center", vertical="center", wrap_text=True),
        )

    thin = Side(style="thin", color="CCCCCC")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    bf   = Font(name="Arial", size=9)
    ALT_FILL = PatternFill("solid", fgColor="EBF3FB")
    WHT_FILL = PatternFill("solid", fgColor="FFFFFF")
    ctr  = Alignment(horizontal="center", vertical="center")
    lft  = Alignment(horizontal="left",   vertical="center")

    def apply_hdr(ws, headers, fill_hex="1F4E79"):
        ws.row_dimensions[1].height = 30
        hs = hdr_style(fill_hex)
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = hs["font"]; c.fill = hs["fill"]
            c.alignment = hs["align"]; c.border = bdr

    def _sanitise(v):
        """Strip illegal XML/openpyxl characters (control chars, null bytes)."""
        if not isinstance(v, str):
            return v
        # Remove null bytes and other control characters that openpyxl rejects
        import re as _re
        return _re.sub('[\\x00-\\x08\\x0b\\x0c\\x0e-\\x1f]', '', v)

    def body_cell(ws, row, col, val, alt_row=False, centre=False):
        c = ws.cell(row=row, column=col, value=_sanitise(val))
        c.font = bf
        c.fill = ALT_FILL if alt_row else WHT_FILL
        c.border = bdr
        c.alignment = ctr if centre else lft
        return c

    def col_w(ws, widths):
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

    def add_autofilter(ws, ncols, hdr_row=1):
        """Add Excel AutoFilter drop-downs to the header row."""
        ws.auto_filter.ref = f"A{hdr_row}:{get_column_letter(ncols)}{hdr_row}"

    # ── Sheet 1: Connections ──────────────────────────────────────────────────
    ws1 = wb.active; ws1.title = "Connections"
    HDR1 = ["Hostname (src)","MAC Address (src)","Source IP",
            "Hostname (dst)","MAC Address (dst)","Destination IP",
            "Protocol","Port","Resource (if HTTP/HTTPS)",
            "Assessed Server Role","Packets"]
    apply_hdr(ws1, HDR1)
    ws1.freeze_panes = "A2"
    add_autofilter(ws1, len(HDR1))

    def hn(ip):  return nodes.get(ip,{}).get("hostname","")
    def fm(ip):
        macs = sorted(nodes.get(ip,{}).get("macs",set()))
        return macs[0] if macs else ""
    def rl(ip):  return nodes.get(ip,{}).get("role","unknown")

    agg = defaultdict(lambda: dict(count=0, resources=set()))
    for r in rows:
        k = (r["src_ip"], r["dst_ip"], r["proto"], r["port"],
             r.get("src_mac",""), r.get("dst_mac",""))
        agg[k]["count"] += 1
        if r.get("resource"): agg[k]["resources"].add(r["resource"])

    for ri, ((src_ip,dst_ip,proto,port,smac,dmac), info) in \
            enumerate(sorted(agg.items()), 2):
        ws1.row_dimensions[ri].height = 15
        alt = (ri % 2 == 0)
        vals = [hn(src_ip), smac or fm(src_ip), src_ip,
                hn(dst_ip), dmac or fm(dst_ip), dst_ip,
                proto, port or "", "; ".join(sorted(info["resources"])),
                rl(dst_ip), info["count"]]
        for ci, v in enumerate(vals, 1):
            body_cell(ws1, ri, ci, v, alt, centre=(ci in (7,8,11)))

    col_w(ws1, [18,18,15,18,18,15,13,7,28,18,10])

    # ── Sheet 2: Node Summary ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Node Summary")
    HDR2 = ["IP Address","Hostname","MAC Address(es)","Subnet","Role",
            "OS Guess (TTL)","TTL Value","Is Private",
            "Protocols Observed","Passive Open Ports",
            "Pentest Flags","Packet Count","Bytes"]
    apply_hdr(ws2, HDR2, "1F4E79")
    ws2.freeze_panes = "A2"
    add_autofilter(ws2, len(HDR2))

    def node_sort(item):
        ip, info = item
        r = {"server":0,"host":1,"client":2,"external":3}.get(info["role"],4)
        try: return (r, int(ipaddress.ip_address(ip.split("/")[0])))
        except: return (r, ip)

    for ri, (ip, info) in enumerate(sorted(nodes.items(), key=node_sort), 2):
        ws2.row_dimensions[ri].height = 15
        alt = (ri%2==0)
        vals = [
            ip,
            info.get("hostname",""),
            ", ".join(sorted(info["macs"])),
            info["subnet"],
            info["role"],
            info.get("os_guess","Unknown"),
            info.get("ttl_val",""),
            "Yes" if info["is_private"] else "No",
            ", ".join(sorted(info["protocols"])),
            ", ".join(str(p) for p in sorted(info["open_ports"])),
            "; ".join(sorted(info.get("flags",set()))),
            info["count"],
            info["bytes"],
        ]
        for ci, v in enumerate(vals, 1):
            c = body_cell(ws2, ri, ci, v, alt, centre=(ci in (6,7,8,12,13)))
            # Highlight flagged rows
            if info.get("flags"):
                if ci == 11:
                    c.font = Font(name="Arial", size=9, color="B85450", bold=True)

    col_w(ws2, [15,18,30,18,10,15,9,10,45,22,28,12,12])

    # ── Sheet 3: Pentest Findings ─────────────────────────────────────────────
    ws3 = wb.create_sheet("Pentest Findings")
    HDR3 = ["Severity","Category","Source","Destination","Detail","Recommendation"]
    apply_hdr(ws3, HDR3, "8B0000")
    ws3.freeze_panes = "A2"
    add_autofilter(ws3, len(HDR3))

    SEV_COLOURS = {"CRITICAL":"FF0000","HIGH":"FF6600",
                   "MEDIUM":"FFB300","LOW":"00AA00","INFO":"888888"}

    sev_order = {"CRITICAL":0,"HIGH":1,"MEDIUM":2,"LOW":3,"INFO":4}
    sorted_findings = sorted(findings, key=lambda f: sev_order.get(f["severity"],5))

    for ri, f in enumerate(sorted_findings, 2):
        ws3.row_dimensions[ri].height = 28
        vals = [f["severity"], f["category"], f["src"], f["dst"],
                f["detail"], f["recommendation"]]
        for ci, v in enumerate(vals, 1):
            c = body_cell(ws3, ri, ci, v, False, centre=(ci==1))
            c.alignment = Alignment(horizontal="center" if ci==1 else "left",
                                    vertical="center", wrap_text=True)
            if ci == 1:
                clr = SEV_COLOURS.get(f["severity"],"888888")
                c.font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
                c.fill  = PatternFill("solid", fgColor=clr)
                c.border = bdr

    col_w(ws3, [12,28,16,16,50,42])

    # ── Sheet 4: Protocol Summary ─────────────────────────────────────────────
    ws4 = wb.create_sheet("Protocol Summary")
    HDR4 = ["Protocol","Connections","Unique Sources","Unique Destinations",
            "Cleartext?","Lateral Movement Risk?"]
    apply_hdr(ws4, HDR4, "1F4E79")
    ws4.freeze_panes = "A2"
    add_autofilter(ws4, len(HDR4))

    proto_stats = defaultdict(lambda: dict(count=0, srcs=set(), dsts=set()))
    for (src_ip,dst_ip,proto,port,smac,dmac), info in agg.items():
        proto_stats[proto]["count"] += info["count"]
        proto_stats[proto]["srcs"].add(src_ip)
        proto_stats[proto]["dsts"].add(dst_ip)

    for ri, (proto, ps) in enumerate(
            sorted(proto_stats.items(), key=lambda x: -x[1]["count"]), 2):
        alt = (ri%2==0)
        vals = [proto, ps["count"], len(ps["srcs"]), len(ps["dsts"]),
                "YES ⚠" if proto in CLEARTEXT_PROTOS else "No",
                "YES 🔴" if proto in LATERAL_PROTOS else "No"]
        for ci, v in enumerate(vals, 1):
            c = body_cell(ws4, ri, ci, v, alt, centre=True)
            if ci==5 and v.startswith("YES"):
                c.font = Font(name="Arial", size=9, bold=True, color="B85450")
            if ci==6 and v.startswith("YES"):
                c.font = Font(name="Arial", size=9, bold=True, color="CC0000")

    col_w(ws4, [16,14,16,20,14,22])

    # ── Sheet 5: Port Inventory ───────────────────────────────────────────────
    ws5 = wb.create_sheet("Port Inventory")
    HDR5 = ["IP","Hostname","Role","OS Guess","Subnet",
            "Passive Open Ports","Well-Known Service Names","Suspicious Ports"]
    apply_hdr(ws5, HDR5, "1F4E79")
    ws5.freeze_panes = "A2"
    add_autofilter(ws5, len(HDR5))

    for ri, (ip, info) in enumerate(sorted(nodes.items(), key=node_sort), 2):
        ports    = sorted(info["open_ports"])
        svc_names = [WELL_KNOWN.get(("TCP",p)) or WELL_KNOWN.get(("UDP",p)) or ""
                     for p in ports]
        sus_ports = [str(p) for p in ports if p in SUSPICIOUS_PORTS]
        alt = (ri%2==0)
        vals = [ip, info.get("hostname",""), info["role"],
                info.get("os_guess","Unknown"), info["subnet"],
                ", ".join(str(p) for p in ports),
                ", ".join(s for s in svc_names if s),
                ", ".join(sus_ports) or ""]
        for ci, v in enumerate(vals, 1):
            c = body_cell(ws5, ri, ci, v, alt, centre=(ci in (3,4)))
            if ci==8 and v:
                c.font = Font(name="Arial", size=9, bold=True, color="B85450")

    col_w(ws5, [15,18,10,14,18,28,30,20])


    # ── Sheet 6: Cleartext Intercepts ────────────────────────────────────────
    ws6 = wb.create_sheet("Cleartext Intercepts")
    HDR6 = ["Protocol","Data Type","Extracted Value","Context / Raw",
            "Source IP","Destination IP","Src Port","Dst Port"]
    apply_hdr(ws6, HDR6, "4A0000")
    ws6.freeze_panes = "A2"
    add_autofilter(ws6, len(HDR6))

    # De-duplicate: same (proto, type, value, src, dst) seen in multiple packets
    seen_ct = set()
    deduped = []
    for h in cleartext_hits:
        key = (h["protocol"], h["type"], h["value"], h["src_ip"], h["dst_ip"])
        if key not in seen_ct:
            seen_ct.add(key)
            deduped.append(h)

    TYPE_SEVERITY = {
        "FTP Password":"CRITICAL", "FTP Username":"HIGH",
        "HTTP Basic Auth (decoded)":"CRITICAL", "HTTP Basic Auth (raw b64)":"HIGH",
        "HTTP Bearer Token":"HIGH", "HTTP Cookie":"MEDIUM",
        "HTTP POST Credential":"CRITICAL", "HTTP JSON Credential":"CRITICAL",
        "Telnet Keystrokes/Data":"HIGH", "Telnet Data":"HIGH",
        "SMTP Auth Command":"HIGH", "SMTP Auth (b64 decoded)":"CRITICAL",
        "POP3 Password":"CRITICAL", "POP3 Username":"HIGH",
        "IMAP Login":"CRITICAL",
        "LDAP Bind Data":"HIGH",
        "SNMP Community String":"MEDIUM",
        "AWS Access Key":"CRITICAL", "Private Key":"CRITICAL",
        "API Key":"HIGH", "Auth Token":"HIGH", "Secret/Key":"HIGH",
    }

    SEV_CLR = {
        "CRITICAL": ("FF0000","FFFFFF"),
        "HIGH":     ("FF6600","FFFFFF"),
        "MEDIUM":   ("FFB300","000000"),
        "LOW":      ("00AA00","FFFFFF"),
    }

    # Sort: CRITICAL first, then by protocol
    def ct_sort(h):
        sev = TYPE_SEVERITY.get(h["type"],"LOW")
        return ({"CRITICAL":0,"HIGH":1,"MEDIUM":2,"LOW":3}.get(sev,4), h["protocol"])

    deduped.sort(key=ct_sort)

    if not deduped:
        ws6.cell(row=2, column=1, value="No cleartext credentials or sensitive data detected.")
    else:
        for ri, h in enumerate(deduped, 2):
            ws6.row_dimensions[ri].height = 20
            sev  = TYPE_SEVERITY.get(h["type"], "LOW")
            bg, fg = SEV_CLR.get(sev, ("FFFFFF","000000"))
            alt_row = (ri % 2 == 0)

            vals = [h["protocol"], h["type"], h["value"], h["context"],
                    h["src_ip"], h["dst_ip"], h["src_port"] or "", h["dst_port"] or ""]
            for ci, v in enumerate(vals, 1):
                c = body_cell(ws6, ri, ci, v, alt_row, centre=(ci in (1,7,8)))
                # Colour the Data Type cell by severity
                if ci == 2:
                    c.font  = Font(name="Arial", size=9, bold=True, color=fg)
                    c.fill  = PatternFill("solid", fgColor=bg)
                    c.border = bdr
                # Highlight the value cell for critical items
                if ci == 3 and sev == "CRITICAL":
                    c.font = Font(name="Arial", size=9, bold=True, color="8B0000")

    col_w(ws6, [14, 28, 45, 45, 15, 15, 9, 9])

    # Add a summary note at the top
    ws6.insert_rows(1)
    ws6.row_dimensions[1].height = 40
    summary = ws6.cell(row=1, column=1,
        value=(f"CLEARTEXT INTERCEPTS — {len(deduped)} unique items captured  "
               f"| CRITICAL: {sum(1 for h in deduped if TYPE_SEVERITY.get(h['type'])=='CRITICAL')}  "
               f"HIGH: {sum(1 for h in deduped if TYPE_SEVERITY.get(h['type'])=='HIGH')}  "
               f"MEDIUM: {sum(1 for h in deduped if TYPE_SEVERITY.get(h['type'])=='MEDIUM')}"))
    summary.font  = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    summary.fill  = PatternFill("solid", fgColor="4A0000")
    summary.alignment = Alignment(horizontal="left", vertical="center",
                                   indent=1, wrap_text=False)
    ws6.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)



    # ── Sheet 7: Banner & Resource Intelligence ───────────────────────────────
    ws7 = wb.create_sheet("Banner Intel")

    # Sub-section colours
    CAT_COLOURS = {
        "Banner":          ("1A3A5C", "FFFFFF"),   # dark navy
        "Resource":        ("1A5C2E", "FFFFFF"),   # dark green
        "Client Software": ("5C3A1A", "FFFFFF"),   # dark brown
    }

    # Build structured sections: Banners, then Resources, then Client Software
    def _banner_sort(h):
        cat_order = {"Banner":0, "Resource":1, "Client Software":2}
        return (cat_order.get(h["category"], 9), h["protocol"], h["server_ip"])

    banner_hits_sorted = sorted(banner_hits, key=_banner_sort)

    HDR7 = ["Category", "Banner / Resource Type", "Server / Host IP",
            "Client IP", "Port", "Protocol", "Value / Detail", "Context"]
    apply_hdr(ws7, HDR7, "1A3A5C")
    ws7.freeze_panes = "A2"
    add_autofilter(ws7, len(HDR7))

    BANNER_TYPE_NOTE = {
        "HTTP Server Header":   "⚠ Version disclosure — update server header suppression",
        "X-Powered-By":         "⚠ Technology disclosure — remove in production",
        "X-Generator":          "⚠ CMS/framework version disclosed",
        "ASP.NET Version":      "⚠ .NET version disclosure — disable X-AspNet-Version header",
        "Via (Proxy)":          "ℹ Proxy/load balancer in path",
        "FTP Banner":           "⚠ Server version in banner — consider suppressing",
        "SMTP Banner":          "⚠ Mail server version disclosed in greeting",
        "SSH Version String":   "⚠ SSH version disclosed — consider 'DebianBanner no'",
        "Telnet Version String":"⚠ Device version in Telnet banner",
        "Telnet Login Banner":  "ℹ Login banner content",
        "SNMP sysDescr":        "⚠ OS/device info in SNMP — consider restricting access",
        "DHCP Vendor Class":    "ℹ Client identifies itself to DHCP server",
        "Runtime Version":      "⚠ Runtime version disclosure",
        "HTTP Request":         "ℹ Commonly requested resource",
        "HTTP User-Agent":      "ℹ Client software / OS fingerprint",
        "DNS Query":            "ℹ Domain being resolved",
        "NTP Server":           "ℹ NTP time source",
        "DHCP Server":          "ℹ DHCP server identity",
        "SMTP EHLO Domain":     "ℹ Client mail domain",
    }

    # De-duplicate banner_hits (already done inside extract_banners, but wb sheet needs it)
    seen_b = set()
    deduped_b = []
    for h in banner_hits_sorted:
        key = (h["category"], h["banner_type"], h["server_ip"], h["value"])
        if key not in seen_b:
            seen_b.add(key)
            deduped_b.append(h)

    if not deduped_b:
        ws7.cell(row=2, column=1, value="No banners or notable resources detected.")
    else:
        current_cat = None
        data_row = 2

        for h in deduped_b:
            cat = h["category"]

            # Insert a category divider row when the category changes
            if cat != current_cat:
                current_cat = cat
                bg_hex, fg_hex = CAT_COLOURS.get(cat, ("444444","FFFFFF"))
                div_cell = ws7.cell(row=data_row, column=1,
                                    value=f"── {cat.upper()} ──")
                div_cell.font  = Font(name="Arial", bold=True, size=10, color=fg_hex)
                div_cell.fill  = PatternFill("solid", fgColor=bg_hex)
                div_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                div_cell.border = bdr
                # Merge across all columns for the divider
                ws7.merge_cells(start_row=data_row, start_column=1,
                                end_row=data_row, end_column=8)
                ws7.row_dimensions[data_row].height = 18
                data_row += 1

            ws7.row_dimensions[data_row].height = 22
            note = BANNER_TYPE_NOTE.get(h["banner_type"], "")
            vals = [
                h["category"],
                h["banner_type"],
                h["server_ip"],
                h["client_ip"],
                h["port"] or "",
                h["protocol"],
                h["value"],
                note or h.get("context",""),
            ]
            for ci, v in enumerate(vals, 1):
                c = body_cell(ws7, data_row, ci, v, (data_row % 2 == 0),
                              centre=(ci in (1,5,6)))
                c.alignment = Alignment(
                    horizontal="center" if ci in (1,5,6) else "left",
                    vertical="center", wrap_text=(ci in (7,8)))
                # Colour-code the Category column
                if ci == 1:
                    bg_hex, fg_hex = CAT_COLOURS.get(cat, ("444444","FFFFFF"))
                    c.font  = Font(name="Arial", size=9, bold=True, color=fg_hex)
                    c.fill  = PatternFill("solid", fgColor=bg_hex)
                    c.border = bdr
                # Highlight ⚠ notes in the Context column
                if ci == 8 and str(v).startswith("⚠"):
                    c.font = Font(name="Arial", size=9, color="B85450", bold=True)

            data_row += 1

    col_w(ws7, [16, 26, 16, 16, 7, 12, 55, 48])

    # Summary header at top
    ws7.insert_rows(1)
    ws7.row_dimensions[1].height = 36
    n_banners   = sum(1 for h in deduped_b if h["category"] == "Banner")
    n_resources = sum(1 for h in deduped_b if h["category"] == "Resource")
    n_clients   = sum(1 for h in deduped_b if h["category"] == "Client Software")
    summ = ws7.cell(row=1, column=1,
        value=(f"BANNER & RESOURCE INTELLIGENCE  |  "
               f"Service Banners: {n_banners}   "
               f"Resources / Queries: {n_resources}   "
               f"Client Software: {n_clients}"))
    summ.font  = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    summ.fill  = PatternFill("solid", fgColor="1A3A5C")
    summ.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws7.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)



    # ── Sheet 8: TLS Session Analysis ────────────────────────────────────────
    ws8 = wb.create_sheet("TLS Sessions")

    HDR8 = [
        "Risk",
        "Client IP", "Server IP", "Port",
        "SNI / Hostname", "ALPN",
        "TLS Version (Negotiated)", "Cipher Suite",
        "Cert Subject", "Cert Issuer", "SANs",
        "Cert Valid From", "Cert Expiry",
        "Key Type", "Key Bits",
        "Handshake Complete", "Alerts",
        "Issues / Findings",
    ]
    apply_hdr(ws8, HDR8, "1B3A5C")
    ws8.freeze_panes = "A2"
    add_autofilter(ws8, len(HDR8))
    ws8.row_dimensions[1].height = 36

    # Severity helpers
    def _tls_risk(sess):
        if sess.get("cert_expired"):        return "CRITICAL"
        if sess.get("weak_cipher"):         return "HIGH"
        if sess.get("weak_version"):        return "HIGH"
        if sess.get("cert_expiring_soon"):  return "MEDIUM"
        if sess.get("issues"):              return "MEDIUM"
        if not sess.get("tls_version"):     return "INFO"
        return "OK"

    TLS_RISK_CLR = {
        "CRITICAL": ("C00000", "FFFFFF"),
        "HIGH":     ("C55A11", "FFFFFF"),
        "MEDIUM":   ("BF8F00", "000000"),
        "INFO":     ("595959", "FFFFFF"),
        "OK":       ("375623", "FFFFFF"),
    }

    # Sort: worst issues first, then by server IP
    def _tls_sort(s):
        order = {"CRITICAL":0,"HIGH":1,"MEDIUM":2,"INFO":3,"OK":4}
        return (order.get(_tls_risk(s), 9), s["server_ip"], s["server_port"])

    tls_sorted = sorted(tls_sessions, key=_tls_sort)

    if not tls_sorted:
        ws8.cell(row=2, column=1,
                 value="No TLS handshakes detected in capture — only encrypted app-data seen, or no TLS traffic present.")
    else:
        for ri, sess in enumerate(tls_sorted, 2):
            ws8.row_dimensions[ri].height = 22
            alt = (ri % 2 == 0)
            risk = _tls_risk(sess)
            risk_bg, risk_fg = TLS_RISK_CLR.get(risk, ("FFFFFF","000000"))

            cs = sess.get("cipher_suite","")
            ver = sess.get("tls_version","")
            expiry = sess.get("cert_not_after","")

            issues_str = "; ".join(sess.get("issues",[])) if sess.get("issues") else ""
            alerts_str = "; ".join(sess.get("alerts",[])) if sess.get("alerts") else ""

            vals = [
                risk,
                sess["client_ip"],
                sess["server_ip"],
                sess["server_port"] or "",
                sess.get("sni",""),
                sess.get("alpn",""),
                ver,
                cs,
                sess.get("cert_subject",""),
                sess.get("cert_issuer",""),
                sess.get("cert_sans",""),
                sess.get("cert_not_before",""),
                expiry,
                sess.get("cert_key_type",""),
                sess.get("cert_key_bits","") or "",
                "Yes" if sess.get("handshake_complete") else "Partial",
                alerts_str,
                issues_str,
            ]

            for ci, v in enumerate(vals, 1):
                centre_cols = {1, 4, 6, 7, 14, 15, 16}
                c = body_cell(ws8, ri, ci, v, alt, centre=(ci in centre_cols))
                c.alignment = Alignment(
                    horizontal="center" if ci in centre_cols else "left",
                    vertical="center",
                    wrap_text=(ci in {8, 11, 18}),
                )

                # Risk column — colour-coded badge
                if ci == 1:
                    c.font  = Font(name="Arial", bold=True, size=9, color=risk_fg)
                    c.fill  = PatternFill("solid", fgColor=risk_bg)
                    c.border = bdr

                # TLS version — red if deprecated
                elif ci == 7 and ver in ("SSLv3","TLS 1.0","TLS 1.1"):
                    c.font = Font(name="Arial", size=9, bold=True, color="C00000")

                # Cipher suite — orange if weak
                elif ci == 8 and sess.get("weak_cipher"):
                    c.font = Font(name="Arial", size=9, bold=True, color="C55A11")

                # Cert expiry — red if expired, amber if soon
                elif ci == 13:
                    if sess.get("cert_expired"):
                        c.font = Font(name="Arial", size=9, bold=True, color="C00000")
                    elif sess.get("cert_expiring_soon"):
                        c.font = Font(name="Arial", size=9, bold=True, color="BF8F00")

                # Key bits — red if weak
                elif ci == 15 and isinstance(v, int):
                    kt = sess.get("cert_key_type","")
                    if ("RSA" in kt and v and v < 2048) or ("ECDSA" in kt and v and v < 256):
                        c.font = Font(name="Arial", size=9, bold=True, color="C00000")

                # Issues — bold red if non-empty
                elif ci == 18 and v:
                    c.font = Font(name="Arial", size=9, bold=True, color="8B0000")

    col_w(ws8, [10, 15, 15, 6, 28, 10, 14, 38, 28, 28, 32, 13, 13, 14, 8, 10, 28, 48])

    # Summary banner at top
    ws8.insert_rows(1)
    ws8.row_dimensions[1].height = 36
    n_ok       = sum(1 for s in tls_sessions if _tls_risk(s) == "OK")
    n_critical = sum(1 for s in tls_sessions if _tls_risk(s) == "CRITICAL")
    n_high     = sum(1 for s in tls_sessions if _tls_risk(s) == "HIGH")
    n_medium   = sum(1 for s in tls_sessions if _tls_risk(s) == "MEDIUM")
    n_info     = sum(1 for s in tls_sessions if _tls_risk(s) == "INFO")
    tls_summ = ws8.cell(row=1, column=1,
        value=(f"TLS SESSION ANALYSIS  |  Total: {len(tls_sessions)}   "
               f"✓ OK: {n_ok}   "
               f"⚠ MEDIUM: {n_medium}   "
               f"▲ HIGH: {n_high}   "
               f"✖ CRITICAL: {n_critical}   "
               f"ℹ Info only: {n_info}"))
    tls_summ.font  = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    tls_summ.fill  = PatternFill("solid", fgColor="1B3A5C")
    tls_summ.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws8.merge_cells(start_row=1, start_column=1, end_row=1, end_column=18)



    # ── Sheet 9: DNS / mDNS Events ────────────────────────────────────────────
    ws9 = wb.create_sheet("DNS Events")

    HDR9 = ["Direction","Protocol","Client IP","Server/Resolver IP",
            "Query Name","Type","Answer IP / Value","Answer Name",
            "TTL (s)","Response Code","Flags"]
    apply_hdr(ws9, HDR9, "1A4A1A")
    ws9.freeze_panes = "A2"
    add_autofilter(ws9, len(HDR9))

    RTYPE_CLR = {"A":"D9F0D3","AAAA":"C6E2FF","CNAME":"FFF2CC",
                 "PTR":"FFE0CC","MX":"F0D9F0","TXT":"F0F0F0",
                 "SRV":"E8D5FF","NS":"D0D0D0","HTTPS":"C6F0FF"}
    RCODE_CLR = {"NXDOMAIN":"FFC0C0","SERVFAIL":"FFB366","REFUSED":"FFD0D0"}

    if not dns_events:
        ws9.cell(row=2, column=1, value="No DNS/mDNS events captured in this PCAP.")
    else:
        # Sort: queries first, then responses; within each by client IP
        def _dns_sort(e):
            return (0 if not e["is_response"] else 1,
                    e.get("client_ip",""),
                    e.get("query_name",""))
        sorted_events = sorted(dns_events, key=_dns_sort)

        for ri, ev in enumerate(sorted_events, 2):
            ws9.row_dimensions[ri].height = 20
            alt = (ri % 2 == 0)
            direction = "← Response" if ev["is_response"] else "→ Query"
            rcode = ev.get("rcode","")

            vals = [
                direction,
                ev.get("proto","DNS"),
                ev.get("client_ip",""),
                ev.get("server_ip",""),
                ev.get("query_name",""),
                ev.get("qtype",""),
                ev.get("answer_val",""),
                ev.get("answer_name",""),
                ev.get("ttl","") or "",
                rcode,
                ev.get("flags_desc",""),
            ]
            for ci, v in enumerate(vals, 1):
                c = body_cell(ws9, ri, ci, v, alt, centre=(ci in {1,2,6,9,10}))
                c.alignment = Alignment(
                    horizontal="center" if ci in {1,2,6,9,10} else "left",
                    vertical="center",
                    wrap_text=(ci in {5,7,8,11}))

                # Colour direction column
                if ci == 1:
                    clr = "1A4A1A" if "Response" in str(v) else "4A1A1A"
                    c.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
                    c.fill = PatternFill("solid", fgColor=clr)

                # Colour record type
                elif ci == 6:
                    qtype_bg = RTYPE_CLR.get(str(v),"FFFFFF")
                    c.fill   = PatternFill("solid", fgColor=qtype_bg)
                    c.font   = Font(name="Arial", size=9, bold=True)

                # Red for NXDOMAIN etc
                elif ci == 10 and rcode in RCODE_CLR:
                    c.fill = PatternFill("solid", fgColor=RCODE_CLR[rcode])
                    c.font = Font(name="Arial", size=9, bold=True, color="8B0000")

    col_w(ws9, [12, 8, 16, 18, 38, 8, 32, 28, 7, 10, 30])

    # Summary banner
    ws9.insert_rows(1)
    ws9.row_dimensions[1].height = 32
    n_queries   = sum(1 for e in dns_events if not e["is_response"])
    n_responses = sum(1 for e in dns_events if e["is_response"])
    n_nxdomain  = sum(1 for e in dns_events if e.get("rcode")=="NXDOMAIN")
    n_mdns      = sum(1 for e in dns_events if e.get("proto")=="mDNS")
    unique_names = len({e["query_name"] for e in dns_events if e.get("query_name")})
    dns_summ = ws9.cell(row=1, column=1,
        value=(f"DNS & mDNS EVENT LOG  |  "
               f"Total: {len(dns_events)}   "
               f"Queries: {n_queries}   Responses: {n_responses}   "
               f"mDNS: {n_mdns}   NXDOMAIN: {n_nxdomain}   "
               f"Unique names queried: {unique_names}"))
    dns_summ.font  = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    dns_summ.fill  = PatternFill("solid", fgColor="1A4A1A")
    dns_summ.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws9.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)


    # ── Sheet 10: Wi-Fi Networks & Clients ────────────────────────────────────
    ws10 = wb.create_sheet("Wi-Fi Networks")
    aps_list     = wifi_data.get("aps", [])
    clients_list = wifi_data.get("clients", [])
    wifi_events  = wifi_data.get("events", [])

    if not aps_list and not clients_list:
        ws10.cell(row=1, column=1,
                  value=("No 802.11 management frames found. "
                         "To capture Wi-Fi data, use monitor mode: "
                         "airmon-ng start wlan0 && tcpdump -i wlan0mon -w wifi.pcap"))
        col_w(ws10, [80])
    else:
        # ── AP table ──────────────────────────────────────────────────────
        HDR_AP = ["BSSID (AP MAC)","SSID","Channel","Encryption",
                  "WPS","Beacons","Probe Resp","Associated Clients","Notes"]
        apply_hdr(ws10, HDR_AP, "1B2A5C")
        ws10.freeze_panes = "A2"
        add_autofilter(ws10, len(HDR_AP))

        ENC_CLR = {"Open":"C00000","WPA":"C55A11","WPA2":"375623",
                   "WPA3":"1F7A1F","WPA2+FT":"375623"}

        for ri, ap in enumerate(aps_list, 2):
            ws10.row_dimensions[ri].height = 22
            alt = (ri % 2 == 0)
            enc = ap.get("enc","Open")
            ssid = ap.get("ssid","<Hidden>") or "<Hidden>"
            notes = []
            if enc == "Open":     notes.append("⚠ Open network — no encryption")
            if ap.get("wps"):     notes.append("⚠ WPS enabled — KRACK/Pixie-Dust risk")
            if not ap.get("ssid"): notes.append("Hidden SSID")

            vals = [
                ap["bssid"],
                ssid,
                ap.get("channel","") or "",
                enc,
                "Yes" if ap.get("wps") else "No",
                ap.get("beacons",0),
                ap.get("probe_responses",0),
                ", ".join(ap.get("clients",[])) or "None seen",
                "; ".join(notes),
            ]
            for ci, v in enumerate(vals, 1):
                c = body_cell(ws10, ri, ci, v, alt, centre=(ci in {3,5,6,7}))
                c.alignment = Alignment(
                    horizontal="center" if ci in {3,5,6,7} else "left",
                    vertical="center", wrap_text=(ci in {8,9}))
                # Colour encryption column
                if ci == 4:
                    enc_bg = ENC_CLR.get(enc.split("+")[0], "595959")
                    c.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
                    c.fill = PatternFill("solid", fgColor=enc_bg)
                # Red for open / WPS
                elif ci == 9 and notes:
                    c.font = Font(name="Arial", size=9, bold=True, color="8B0000")

        col_w(ws10, [20, 28, 8, 10, 6, 9, 10, 45, 45])

        # ── Clients section ───────────────────────────────────────────────
        client_start = len(aps_list) + 4
        ws10.cell(row=client_start-1, column=1, value="Wi-Fi Clients / Stations")
        hdr_row = ws10.row_dimensions[client_start-1]
        hdr_row.height = 24
        c_hdr = ws10.cell(row=client_start-1, column=1)
        c_hdr.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        c_hdr.fill = PatternFill("solid", fgColor="3A1A5C")
        ws10.merge_cells(start_row=client_start-1, start_column=1,
                         end_row=client_start-1, end_column=4)

        HDR_CL = ["Client MAC","Associated BSSID","Associated SSID","Probed SSIDs"]
        for ci, h in enumerate(HDR_CL, 1):
            c = ws10.cell(row=client_start, column=ci, value=h)
            c.font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
            c.fill  = PatternFill("solid", fgColor="3A1A5C")
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = bdr

        for ri, cl in enumerate(clients_list, client_start+1):
            ws10.row_dimensions[ri].height = 20
            alt = (ri % 2 == 0)
            vals = [
                cl["mac"],
                cl.get("associated_bssid",""),
                cl.get("assoc_ssid",""),
                ", ".join(sorted(cl.get("probed_ssids", set()))) or "(wildcard only)",
            ]
            for ci, v in enumerate(vals, 1):
                body_cell(ws10, ri, ci, v, alt)

        # ── Events section ────────────────────────────────────────────────
        if wifi_events:
            evt_start = client_start + len(clients_list) + 4
            ws10.cell(row=evt_start-1, column=1, value="802.11 Management Frame Events")
            c_hdr2 = ws10.cell(row=evt_start-1, column=1)
            c_hdr2.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
            c_hdr2.fill = PatternFill("solid", fgColor="1A3A5C")
            ws10.merge_cells(start_row=evt_start-1, start_column=1,
                             end_row=evt_start-1, end_column=6)

            HDR_EV = ["Frame Type","Source MAC","Destination MAC","BSSID","SSID","Detail"]
            for ci, h in enumerate(HDR_EV, 1):
                c = ws10.cell(row=evt_start, column=ci, value=h)
                c.font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
                c.fill  = PatternFill("solid", fgColor="1A3A5C")
                c.alignment = Alignment(horizontal="left", vertical="center")
                c.border = bdr

            EVT_CLR = {
                "Beacon":             "EBF3FB",
                "Probe Request":      "FFF2CC",
                "Probe Response":     "E2EFDA",
                "Association Request":"D9D9FF",
                "Deauthentication":   "FFD0D0",
                "Disassociation":     "FFDCC0",
            }
            # Limit to 500 events in sheet (can be huge)
            shown_events = wifi_events[:500]
            for ri, ev in enumerate(shown_events, evt_start+1):
                ws10.row_dimensions[ri].height = 18
                alt = (ri % 2 == 0)
                ft = ev.get("frame_type","")
                vals = [
                    ft,
                    ev.get("src_mac",""),
                    ev.get("dst_mac",""),
                    ev.get("bssid",""),
                    ev.get("ssid",""),
                    ev.get("detail",""),
                ]
                row_fill = EVT_CLR.get(ft, "FFFFFF")
                for ci, v in enumerate(vals, 1):
                    c = body_cell(ws10, ri, ci, v, False)
                    c.fill = PatternFill("solid", fgColor=row_fill)
                    c.alignment = Alignment(horizontal="left", vertical="center")
                    if ft in ("Deauthentication","Disassociation") and ci == 1:
                        c.font = Font(name="Arial", size=9, bold=True, color="C00000")

        # Summary banner (insert at top)
        ws10.insert_rows(1)
        ws10.row_dimensions[1].height = 32
        n_open = sum(1 for a in aps_list if a.get("enc","Open") == "Open")
        n_wps  = sum(1 for a in aps_list if a.get("wps"))
        n_deauth = sum(1 for e in wifi_events if e.get("frame_type") == "Deauthentication")
        wifi_summ = ws10.cell(row=1, column=1,
            value=(f"Wi-Fi NETWORK SURVEY  |  "
                   f"APs: {len(aps_list)}   "
                   f"Clients: {len(clients_list)}   "
                   f"Open networks: {n_open}   "
                   f"WPS enabled: {n_wps}   "
                   f"Deauth frames: {n_deauth}   "
                   f"Total mgmt events: {len(wifi_events)}"))
        wifi_summ.font  = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        wifi_summ.fill  = PatternFill("solid", fgColor="1B2A5C")
        wifi_summ.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws10.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)


    # ── Sheet 11: Certificates (TLS + EAP-TLS/802.1X) ────────────────────────
    certificates = certificates or []
    ws11 = wb.create_sheet("Certificates")
    if not certificates:
        ws11.cell(row=1, column=1,
                  value="No X.509 certificates observed (no TLS handshakes or "
                        "EAP-TLS/802.1X exchanges with certificate exchange found).")
        col_w(ws11, [100])
    else:
        HDR11 = ["Source","Context","SNI","Subject CN","Issuer CN","SANs",
                 "Not Before","Not After","Expired?","Key Type","Key Bits",
                 "SHA-256 Fingerprint"]
        apply_hdr(ws11, HDR11, "5C1A1A")
        ws11.freeze_panes = "A2"
        add_autofilter(ws11, len(HDR11))

        for ri, cert in enumerate(certificates, 2):
            ws11.row_dimensions[ri].height = 18
            alt = (ri % 2 == 0)
            kt, kb = cert.get("key_type",""), cert.get("key_bits",0)
            weak_key = (("RSA" in kt and kb and kb < 2048) or
                        ("ECDSA" in kt and kb and kb < 256))
            if cert.get("expired"):
                status = "EXPIRED"
            elif cert.get("expiring_soon"):
                status = "Expiring soon"
            else:
                status = "OK"
            vals = [
                cert.get("source",""),
                cert.get("context",""),
                cert.get("sni",""),
                cert.get("subject",""),
                cert.get("issuer",""),
                cert.get("sans",""),
                cert.get("not_before",""),
                cert.get("not_after",""),
                status,
                kt,
                kb or "",
                cert.get("fingerprint_sha256",""),
            ]
            for ci, v in enumerate(vals, 1):
                c = body_cell(ws11, ri, ci, v, alt, centre=(ci in (1,7,8,9,10,11)))
                if ci == 1 and v == "EAP-TLS":
                    c.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
                    c.fill = PatternFill("solid", fgColor="5C1A1A")
                elif ci == 9:
                    if status == "EXPIRED":
                        c.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
                        c.fill = PatternFill("solid", fgColor="C00000")
                    elif status == "Expiring soon":
                        c.font = Font(name="Arial", size=9, bold=True, color="000000")
                        c.fill = PatternFill("solid", fgColor="FFD966")
                elif ci == 10 and weak_key:
                    c.font = Font(name="Arial", size=9, bold=True, color="C00000")

        col_w(ws11, [9, 40, 28, 26, 26, 40, 11, 11, 12, 10, 9, 66])

        n_expired = sum(1 for c in certificates if c.get("expired"))
        n_eaptls  = sum(1 for c in certificates if c["source"] == "EAP-TLS")
        ws11.insert_rows(1)
        ws11.row_dimensions[1].height = 28
        cert_summ = ws11.cell(row=1, column=1,
            value=(f"CERTIFICATES  |  Total: {len(certificates)}   "
                   f"EAP-TLS/802.1X: {n_eaptls}   Expired: {n_expired}"))
        cert_summ.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        cert_summ.fill = PatternFill("solid", fgColor="5C1A1A")
        cert_summ.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws11.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HDR11))


    wb.save(output_path)
    return True

